from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import transaction
from datetime import timedelta
from django.core.exceptions import ValidationError
from .forms import RefillForm, UploadExcelForm
from .models import Refill, Facility
import pandas as pd
from django.utils import timezone
import openpyxl
from django.http import HttpResponse
from django.conf import settings
from django.db.models import F, Q
from django.core.paginator import Paginator
from openpyxl import Workbook
from datetime import datetime
from .forms import UploadExcelForm
from django.contrib import messages
from openpyxl.styles import Font
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout


# views.py



def signin_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('pw')

        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Login successful!')
            return redirect('dashboard')
    return render(request, 'signin.html')



def logout_view(request):
    logout(request)
    messages.success(request, ' Logout successfully')
    return redirect('login')


from io import BytesIO

VALID_REFILL_MONTHS = [0.5, 1, 2, 2.8, 3, 4, 5, 6]

def import_refills_from_excel(file):
    """
    Import refill data from Excel containing multiple facilities.
    Deletes old data per facility before inserting new rows.
    """

    if file.size > 1073741824:  # 1GB
        raise ValidationError("File size exceeds the maximum allowed limit of 1GB.")

    file.seek(0)
    df = pd.read_excel(file)

    # ================= REQUIRED COLUMNS =================
    required_columns = [
        'Unique Id',
        'Last Pickup Date (yyyy-mm-dd)',
        'Months of ARV Refill',
        'Current ART Regimen',
        'Case Manager',
        'Sex',
        'Current ART Status',
        'Facility Name',
        'ART Start Date (yyyy-mm-dd)',
        'Date of Viral Load Sample Collection (yyyy-mm-dd)',
    ]

    for col in required_columns:
        if col not in df.columns:
            raise ValidationError(f"Missing column: {col}")

    # Only Active / Active Restart
    df = df[df['Current ART Status'].isin(['Active', 'Active Restart'])]
    if df.empty:
        raise ValidationError("No Active or Active Restart patients found.")

    # Clean facility names
    df['Facility Name'] = df['Facility Name'].astype(str).str.strip()
    facilities = {f.name.strip(): f for f in Facility.objects.filter(
        name__in=df['Facility Name'].unique()
    )}
    missing_facilities = set(df['Facility Name'].unique()) - set(facilities.keys())
    if missing_facilities:
        raise ValidationError(
            f"These facilities do not exist in the system: {', '.join(missing_facilities)}"
        )

    # ================= VALIDATE ROWS =================
    validated_rows = []
    for _, row in df.iterrows():
        unique_id = row['Unique Id']

        # Last pickup date
        try:
            last_pickup_date = pd.to_datetime(
                row['Last Pickup Date (yyyy-mm-dd)']
            ).date()
        except Exception:
            raise ValidationError(
                f"Invalid Last Pickup Date format for Unique Id {unique_id}"
            )

        # Refill months
        try:
            months = float(row['Months of ARV Refill'])
        except Exception:
            raise ValidationError(
                f"Invalid Months of ARV Refill for Unique Id {unique_id}"
            )

        if months not in VALID_REFILL_MONTHS:
            raise ValidationError(
                f"Invalid refill duration {months} months "
                f"for Unique Id {unique_id}. Allowed values: {VALID_REFILL_MONTHS}"
            )

        # Facility
        facility_obj = facilities[row['Facility Name']]

        # Appointments
        refill_days = months * 30
        next_appointment = last_pickup_date + timedelta(days=refill_days)

        # Optional VL fields
        art_start_date = pd.to_datetime(
            row['ART Start Date (yyyy-mm-dd)']).date() if pd.notnull(
            row['ART Start Date (yyyy-mm-dd)']) else None
        vl_sample_collection_date = pd.to_datetime(
            row['Date of Viral Load Sample Collection (yyyy-mm-dd)']).date() if pd.notnull(
            row['Date of Viral Load Sample Collection (yyyy-mm-dd)']) else None

        validated_rows.append(
            Refill(
                facility=facility_obj,
                unique_id=unique_id,
                last_pickup_date=last_pickup_date,
                months_of_refill_days=months,
                next_appointment=next_appointment,
                current_regimen=str(row['Current ART Regimen']).strip(),
                case_manager=str(row['Case Manager']).strip(),
                sex=str(row['Sex']).strip(),
                current_art_status=row['Current ART Status'].strip(),
                art_start_date=art_start_date,
                vl_sample_collection_date=vl_sample_collection_date,
            )
        )

    # ================= DELETE OLD AND INSERT NEW =================
    facility_ids = {obj.facility.id for obj in validated_rows}

    with transaction.atomic():
        for facility_id in facility_ids:
            Refill.objects.filter(facility_id=facility_id).delete()

        Refill.objects.bulk_create(validated_rows, batch_size=1000)

    return len(validated_rows)







    # Normalize facility names
    
def upload_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        excel_file = request.FILES.get('file')

        if excel_file and excel_file.size > 1073741824:  # 1GB
            return render(request, 'upload.html', {
                'form': form,
                'error': "File size exceeds the 1GB limit."
            })

        if form.is_valid():
            try:
                import_refills_from_excel(excel_file)
                return redirect('refill_list')
            except ValidationError as e:
                return render(request, 'upload.html', {
                    'form': form,
                    'error': str(e)
                })
        else:
            return render(request, 'upload.html', {'form': form})
    else:
        form = UploadExcelForm()

    return render(request, 'upload.html', {'form': form})

def upload_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)

        if not request.FILES:
            messages.error(request, "No file was uploaded.")
            return redirect('upload_excel')

        if form.is_valid():
            excel_file = form.cleaned_data['file']

            try:
                import_refills_from_excel(excel_file)
                messages.success(request, "Excel uploaded and processed successfully!")
                return redirect('upload_excel')

            except Exception as e:
                messages.error(request, f"Upload failed: {str(e)}")
                return redirect('upload_excel')
        else:
            messages.error(request, "Form validation failed.")
            print("FORM ERRORS:", form.errors)

    else:
        form = UploadExcelForm()

    return render(request, 'upload.html', {'form': form})

# ================================
# DASHBOARD
# ================================



def dashboard(request):
    today = timezone.now().date()
    week_end = today + timedelta(days=7)

    month_start = today.replace(day=1)
    month_end = (
        (today.replace(month=today.month+1, day=1) - timedelta(days=1))
        if today.month != 12
        else (today.replace(year=today.year+1, month=1, day=1) - timedelta(days=1))
    )

    facility_id = request.GET.get("facility")
    facilities = Facility.objects.all()

    refills = Refill.objects.filter(current_art_status__in=['Active','Active Restart'])
    selected_facility_instance = None
    if facility_id:
        refills = refills.filter(facility_id=facility_id)
        selected_facility_instance = Facility.objects.filter(id=facility_id).first()

    # Monthly Missed
    monthly_missed_total = refills.filter(
        next_appointment__year=today.year,
        next_appointment__month=today.month,
        next_appointment__lt=today
    ).filter(Q(last_pickup_date__lt=F('next_appointment')) | Q(last_pickup_date__isnull=True)).count()

    # IIT ≥28 days
    iit_total = sum(
        1 for r in refills.filter(next_appointment__lt=today)
        if r.next_appointment and (today - r.next_appointment).days >= 28
    )

    # ====================== VL Quarterly Coverage ======================
    if today.month in [1,2,3]:
        current_quarter = "Q1"
    elif today.month in [4,5,6]:
        current_quarter = "Q2"
    elif today.month in [7,8,9]:
        current_quarter = "Q3"
    else:
        current_quarter = "Q4"

    # Calculate VL eligibility and numerator/denominator
    quarter_start_month = {"Q1":1, "Q2":4, "Q3":7, "Q4":10}[current_quarter]
    quarter_start = timezone.datetime(today.year, quarter_start_month, 1).date()
    quarter_end = (
        (timezone.datetime(today.year, quarter_start_month+3, 1).date() - timedelta(days=1))
        if current_quarter in ["Q1","Q2","Q3"]
        else timezone.datetime(today.year+1, 1, 1).date() - timedelta(days=1)
    )

    eligible_clients = []
    numerator_count = 0

    quarter_refills = refills.filter(art_start_date__lte=quarter_end)

    for r in quarter_refills:
        if r.is_vl_eligible:
            eligible_clients.append(r)
        # Count if VL collected in this quarter
        if r.vl_sample_collection_date and quarter_start <= r.vl_sample_collection_date <= quarter_end:
            numerator_count += 1

    denominator_count = len(eligible_clients)
    vl_coverage = round((numerator_count / denominator_count * 100), 1) if denominator_count > 0 else 0

    coverage_data = {
        "denominator": denominator_count,
        "numerator": numerator_count,
        "coverage": vl_coverage
    }

    # ====================== VL Suppression ======================
    vl_with_result = [r for r in refills if r.vl_result is not None]
    vl_suppressed = sum(1 for r in vl_with_result if r.is_suppressed)
    vl_numerator = len(vl_with_result)
    vl_suppression_rate = round((vl_suppressed / vl_numerator * 100), 1) if vl_numerator else 0

    # ====================== CONTEXT ======================
    context = {
        "facilities": facilities,
        "selected_facility": facility_id,
        "daily_expected": refills.filter(next_appointment=today),
        "daily_refills": refills.filter(last_pickup_date=today),
        "weekly_expected": refills.filter(next_appointment__range=[today, week_end]),
        "weekly_refills": refills.filter(last_pickup_date__range=[today, week_end]),
        "monthly_expected": refills.filter(next_appointment__year=today.year, next_appointment__month=today.month),
        "monthly_refills": refills.filter(last_pickup_date__year=today.year, last_pickup_date__month=today.month),
        "monthly_missed_total": monthly_missed_total,
        "iit_total": iit_total,
        "vl_denominator": coverage_data["denominator"],
        "vl_numerator": coverage_data["numerator"],
        "vl_coverage": coverage_data["coverage"],
        "vl_suppressed": vl_suppressed,
        "vl_suppression_rate": vl_suppression_rate,
        "current_year": today.year,
        "current_quarter": current_quarter,
        "today": today,
    }

    return render(request, "dashboard.html", context)


# ================================
# CRUD VIEWS
# ================================













def refill_list(request):
    today = timezone.now().date()
    week_end = today + timedelta(days=7)

    # =================== GET FILTERS ===================
    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    search_unique_id = request.GET.get("search_unique_id")

    # =================== SAFE DATE PARSING ===================
    start_date_obj = None
    end_date_obj = None

    if start_date and start_date != "None":
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            start_date_obj = None

    if end_date and end_date != "None":
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            end_date_obj = None

    # =================== LOAD DATA ===================
    facilities = Facility.objects.all()
    refills = Refill.objects.all()

    # =================== APPLY FILTERS ===================
    if facility_id and facility_id != "None":
        try:
            refills = refills.filter(facility_id=int(facility_id))
        except ValueError:
            pass

    if selected_case_manager and selected_case_manager != "None":
        refills = refills.filter(case_manager=selected_case_manager)

    if start_date_obj:
        refills = refills.filter(next_appointment__gte=start_date_obj)
    if end_date_obj:
        refills = refills.filter(next_appointment__lte=end_date_obj)

    if search_unique_id and search_unique_id.strip():
        refills = refills.filter(unique_id__icontains=search_unique_id.strip())

    # =================== CASE MANAGER LIST ===================
    case_managers_qs = (
        Refill.objects.exclude(case_manager__isnull=True)
        .exclude(case_manager__exact="")
        .values_list("case_manager", flat=True)
        .distinct()
    )
    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm and cm.strip()})

    # =================== CALCULATE DAYS MISSED & MISSED APPOINTMENT ===================
    for refill in refills:
        if refill.next_appointment and refill.next_appointment < today:
            refill.days_missed = (today - refill.next_appointment).days
            refill.missed_appointment = True
        else:
            refill.days_missed = 0
            refill.missed_appointment = False

        # =================== RISK PREDICTION ===================
        score = 0
        high_risk_keywords = [
            "transport","money","no money","travel","forgot","busy","work","distance",
            "sick","hospital","admitted","defaulted","stopped","side effect"
        ]
        medium_risk_keywords = [
            "delay","reschedule","family issue","school","appointment clash","funeral","religious"
        ]

        if refill.missed_appointment: score += 40
        if refill.days_missed > 30: score += 25
        elif refill.days_missed > 7: score += 15
        if refill.remark:
            remark_lower = refill.remark.lower()
            for word in high_risk_keywords:
                if word in remark_lower: score += 20
            for word in medium_risk_keywords:
                if word in remark_lower: score += 10
        if refill.current_art_status == "Inactive": score += 30
        elif refill.current_art_status == "Active Restart": score += 20
        refill.prediction_probability = min(score, 100)

    # =================== GROUP BY PERIOD ===================
    daily_expected = refills.filter(next_appointment=today)
    weekly_expected = refills.filter(next_appointment__range=[today, week_end])
    monthly_expected = refills.filter(
        next_appointment__year=today.year,
        next_appointment__month=today.month
    )

    daily_page = Paginator(daily_expected.order_by("next_appointment"), 10)
    weekly_page = Paginator(weekly_expected.order_by("next_appointment"), 10)
    monthly_page = Paginator(monthly_expected.order_by("next_appointment"), 10)

    daily_number = request.GET.get("daily_page")
    weekly_number = request.GET.get("weekly_page")
    monthly_number = request.GET.get("monthly_page")

    # =================== EXPORT EXCEL ===================
    if "download" in request.GET:
        return export_refills_to_excel(refills)

    # =================== CONTEXT ===================
    context = {
        "facilities": facilities,
        "selected_facility": facility_id,
        "case_managers": case_managers,
        "selected_case_manager": selected_case_manager,
        "today": today,
        "selected_start_date": start_date,
        "selected_end_date": end_date,
        "search_unique_id": search_unique_id,
        "periods": [
            {"name": "Daily", "page_obj": daily_page.get_page(daily_number)},
            {"name": "Weekly", "page_obj": weekly_page.get_page(weekly_number)},
            {"name": "Monthly", "page_obj": monthly_page.get_page(monthly_number)},
        ],
    }

    return render(request, "refill_list.html", context)


# =================== EXCEL EXPORT ===================
def export_refills_to_excel(refills):
    today = timezone.now().date()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expected Refills Data"

    headers = [
        'Unique ID','Facility','Sex','Current Regimen','Case Manager',
        'Last Pickup','Next Appointment','Days Missed','VL Eligibility Status'
    ]
    ws.append(headers)

    for refill in refills:
        # Calculate dates
        refill.calculate_dates()

        # Days missed
        days_missed = (timezone.now().date() - refill.next_appointment).days if refill.next_appointment and refill.next_appointment < timezone.now().date() else 0

        # VL status
        vl_status = "Eligible" if refill.is_vl_eligible else "Not Eligible"

        row = [
            refill.unique_id,
            refill.facility.name if refill.facility else "",
            refill.sex,
            refill.current_regimen,
            refill.case_manager or "",
            refill.last_pickup_date.strftime("%Y-%m-%d") if refill.last_pickup_date else "Never Picked",
            refill.next_appointment.strftime("%Y-%m-%d") if refill.next_appointment else "",
            days_missed,
            vl_status
        ]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Expected_Refills_{today}.xlsx"'
    wb.save(response)
    return response



def refill_create(request, unique_id=None):
    if unique_id:
        # Fetch the refill by unique_id if passed
        refill = get_object_or_404(Refill, unique_id=unique_id)
    else:
        refill = None  # New refill if no unique_id is passed

    # Get today's date, ensuring it's a datetime.date object
    today = timezone.now().date()  # `today` is a datetime.date object
    
    if refill:
        # Ensure that next_appointment is a datetime.date (not a datetime.datetime)
        if isinstance(refill.next_appointment, datetime):  # Check if it's a datetime object
            refill_next_appointment = refill.next_appointment.date()  # Get only the date part
        else:
            refill_next_appointment = refill.next_appointment  # It's already a datetime.date

        # Example comparison (this is just for illustration, customize based on your logic)
        if refill_next_appointment < today:
            # Logic if the refill's next appointment is in the past
            print("This refill's next appointment is in the past.")

    if request.method == 'POST':
        if refill:
            form = RefillForm(request.POST, instance=refill)  # Edit the existing refill
        else:
            form = RefillForm(request.POST)  # Create a new refill

        if form.is_valid():
            form.save()
            return redirect('daily_refill_list')  # After saving, redirect to the daily refill list

    else:
        if refill:
            form = RefillForm(instance=refill)  # Pre-fill the form if editing
        else:
            form = RefillForm()  # Empty form for new refill

    return render(request, 'refill_form.html', {'form': form})




def refill_update(request, pk):
    """
    Update an existing refill entry and auto-recalculate next appointment.
    """
    refill = get_object_or_404(Refill, pk=pk)
    form = RefillForm(request.POST or None, instance=refill)

    if form.is_valid():
        refill = form.save(commit=False)

        # Auto recalculate next appointment
        refill.next_appointment = (
            refill.last_pickup_date +
            timedelta(days=refill.months_of_refill_days)
        )

        refill.save()
        return redirect('refill_list')

    return render(request, "refill_form.html", {"form": form})












def refill_add_or_update(request, pk=None):
    today = timezone.now().date()

    # If editing an existing refill, get it by pk
    if pk:
        refill = get_object_or_404(Refill, pk=pk)
    else:
        refill = Refill()
        refill.calculate_dates()  # Pre-calculate next_appointment and IIT for new record

    if request.method == "POST":
        form = RefillForm(request.POST, instance=refill)
        if form.is_valid():
            form.save()
            return redirect("refill_list")
    else:
        form = RefillForm(instance=refill)

    return render(
        request,
        "refill_form.html",
        {"form": form, "today": today}
    )











def track_refills(request):
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)

    # ================== FILTERS ==================
    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Safe date parsing
    start_date_obj = None
    end_date_obj = None
    if start_date and start_date != "None":
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            start_date_obj = None
    if end_date and end_date != "None":
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            end_date_obj = None

    facilities = Facility.objects.all()
    refills = Refill.objects.all()

    # Apply filters
    if facility_id and facility_id != "None":
        try:
            refills = refills.filter(facility_id=int(facility_id))
        except ValueError:
            pass
    if selected_case_manager and selected_case_manager != "None":
        refills = refills.filter(case_manager=selected_case_manager)
    if start_date_obj:
        refills = refills.filter(last_pickup_date__gte=start_date_obj)
    if end_date_obj:
        refills = refills.filter(last_pickup_date__lte=end_date_obj)

    # Calculate next appointment & days missed
    for refill in refills:
        refill.calculate_dates()
        if refill.next_appointment and refill.next_appointment < today:
            refill.days_missed = (today - refill.next_appointment).days
            refill.missed_appointment = True
        else:
            refill.days_missed = 0
            refill.missed_appointment = False

    # Group by period
    daily_qs = refills.filter(last_pickup_date=today).order_by('-last_pickup_date')
    weekly_qs = refills.filter(last_pickup_date__gte=start_of_week).order_by('-last_pickup_date')
    monthly_qs = refills.filter(last_pickup_date__gte=start_of_month).order_by('-last_pickup_date')

    # Pagination
    daily_paginator = Paginator(daily_qs, 10)
    weekly_paginator = Paginator(weekly_qs, 10)
    monthly_paginator = Paginator(monthly_qs, 10)

    daily_refills = daily_paginator.get_page(request.GET.get("daily_page"))
    weekly_refills = weekly_paginator.get_page(request.GET.get("weekly_page"))
    monthly_refills = monthly_paginator.get_page(request.GET.get("monthly_page"))

    # Case managers list for filter dropdown
    case_managers_qs = (
        Refill.objects.exclude(case_manager__isnull=True)
        .exclude(case_manager__exact="")
        .values_list("case_manager", flat=True)
        .distinct()
    )
    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm and cm.strip()})

    periods = [
        ('Daily', daily_refills),
        ('Weekly', weekly_refills),
        ('Monthly', monthly_refills),
    ]

    # Excel export
    if 'download' in request.GET:
        return export_track_refills_to_excel(refills)

    context = {
        "facilities": facilities,
        "selected_facility": facility_id,
        "case_managers": case_managers,
        "selected_case_manager": selected_case_manager,
        "today": today,
        "selected_start_date": start_date,
        "selected_end_date": end_date,
        "periods": periods,
    }

    return render(request, "track_refills.html", context)


def export_track_refills_to_excel(refills):
    today = timezone.now().date()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Track Refills Data"

    headers = [
        'Unique ID', 'Facility', 'Last Pickup Date', 'Refill Days', 'Sex',
        'Current Regimen', 'Case Manager', 'Next Appointment',
        'Days Missed', 'VL Eligibility Status'
    ]
    ws.append(headers)

    for refill in refills:
        refill.calculate_dates()
        next_appointment = refill.next_appointment.strftime("%Y-%m-%d") if refill.next_appointment else ""
        last_pickup = refill.last_pickup_date.strftime("%Y-%m-%d") if refill.last_pickup_date else "Never Picked"
        days_missed = (timezone.now().date() - refill.next_appointment).days if refill.next_appointment and refill.next_appointment < timezone.now().date() else 0
        vl_status = "Eligible" if refill.is_vl_eligible else "Not Eligible"

        row = [
            refill.unique_id,
            refill.facility.name if refill.facility else "",
            last_pickup,
            refill.months_of_refill_days,
            refill.sex,
            refill.current_regimen,
            refill.case_manager or "",
            next_appointment,
            days_missed,
            vl_status
        ]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Track_Refills_{today}.xlsx"'
    wb.save(response)
    return response




def export_track_refills_to_excel(refills):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Track Refills Data"

    headers = ['Unique ID', 'Facility', 'Last Pickup Date', 'Refill Days', 'Sex', 'Current Regimen', 'Case Manager', 'Next Appointment']
    ws.append(headers)

    for refill in refills:
        row = [
            refill.unique_id,
            refill.facility.name if refill.facility else "",
            refill.last_pickup_date,
            refill.months_of_refill_days,
            refill.sex,
            refill.current_regimen,
            refill.case_manager or "",
            refill.next_appointment,
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="track_refills.xlsx"'
    wb.save(response)
    return response












def daily_refill_list(request):
    today = timezone.now().date()

    # ================= FACILITIES =================
    facility_id = request.GET.get("facility")
    facilities = Facility.objects.all()

    # ================= CASE MANAGERS =================
    case_managers_qs = (
        Refill.objects.exclude(case_manager__isnull=True)
        .exclude(case_manager__exact="")
        .values_list("case_manager", flat=True)
        .distinct()
    )
    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm.strip()})
    selected_case_manager = request.GET.get("case_manager")

    # ================= DAILY REFILLS =================
    refills = Refill.objects.filter(next_appointment=today).order_by('unique_id')

    if facility_id:
        refills = refills.filter(facility_id=facility_id)

    if selected_case_manager:
        refills = refills.filter(case_manager=selected_case_manager)

    context = {
        "facilities": facilities,
        "selected_facility": facility_id,
        "case_managers": case_managers,
        "selected_case_manager": selected_case_manager,
        "today": today,  # for overdue highlighting
        "refills": refills,
    }

    return render(request, "daily_refill_list.html", context)










def missed_refills(request):
    today = timezone.now().date()

    # ================= GET FILTER PARAMETERS =================
    facility_id = request.GET.get("facility")
    case_manager = request.GET.get("case_manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    search_unique_id = request.GET.get("search_unique_id")

    # ================= BASE QUERYSET =================
    refills = Refill.objects.filter(
        current_art_status__in=["Active", "Active Restart"]
    ).select_related("facility")

    # ================= FILTERS =================
    if facility_id:
        try:
            refills = refills.filter(facility_id=int(facility_id))
        except ValueError:
            pass

    if case_manager:
        refills = refills.filter(case_manager__iexact=case_manager.strip())

    if search_unique_id:
        refills = refills.filter(unique_id__icontains=search_unique_id)

    # ================= DATE FILTER =================
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            refills = refills.filter(next_appointment__gte=start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            refills = refills.filter(next_appointment__lte=end_date_obj)
        except ValueError:
            pass

    # ================= MISSED REFILLS LOGIC =================
    missed_list = refills.filter(next_appointment__lt=today).filter(
        Q(last_pickup_date__lt=F("next_appointment")) |
        Q(last_pickup_date__isnull=True)
    ).order_by("next_appointment")

    # ================= CALCULATE DAYS MISSED AND IIT STATUS =================
    for refill in missed_list:
        if refill.next_appointment:
            days_missed = (today - refill.next_appointment).days
            refill.days_missed = days_missed

            iit_date = refill.next_appointment + timedelta(days=28)
            days_to_iit = (iit_date - today).days

            if days_missed >= 28:
                refill.iit_status = "IIT"
            elif days_missed > 0:
                refill.iit_status = f"{days_to_iit} days to IIT"
            else:
                refill.iit_status = "0"
        else:
            refill.days_missed = 0
            refill.iit_status = "0"

    total_missed = missed_list.count()

    # ================= EXPORT TO EXCEL =================
    if request.GET.get("export") == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from django.http import HttpResponse

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Missed Refills"

        # Add VL Status to headers
        headers = [
            "Unique ID",
            "Case Manager",
            "Facility",
            "Last Pickup",
            "Next Appointment",
            "Days Missed",
            "IIT Status",
            "VL Eligibility Status",  # NEW COLUMN
        ]
        worksheet.append(headers)

        # Make header bold
        for col in range(1, len(headers) + 1):
            worksheet.cell(row=1, column=col).font = Font(bold=True)

        # Append rows
        for refill in missed_list:
            worksheet.append([
                refill.unique_id,
                refill.case_manager or "",
                refill.facility.name if refill.facility else "",
                refill.last_pickup_date.strftime("%Y-%m-%d") if refill.last_pickup_date else "",
                refill.next_appointment.strftime("%Y-%m-%d") if refill.next_appointment else "",
                getattr(refill, "days_missed", 0),
                getattr(refill, "iit_status", ""),
                getattr(refill, "vl_status", "N/A"),  # Use property
            ])

        # Adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 4

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="missed_refills.xlsx"'
        workbook.save(response)
        return response

    # ================= PAGINATION =================
    paginator = Paginator(missed_list, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ================= UNIQUE CASE MANAGERS =================
    case_managers_qs = (
        Refill.objects.exclude(case_manager__isnull=True)
        .exclude(case_manager__exact="")
        .values_list("case_manager", flat=True)
        .distinct()
    )
    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm.strip()})

    # ================= QUERY PARAMS FOR PAGINATION =================
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')

    context = {
        "page_obj": page_obj,
        "today": today,
        "total_missed": total_missed,
        "facilities": Facility.objects.all(),
        "case_managers": case_managers,
        "selected_facility": facility_id,
        "selected_case_manager": case_manager,
        "selected_start_date": start_date,
        "selected_end_date": end_date,
        "search_unique_id": search_unique_id,
        "query_params": query_params.urlencode(),  # safe for template links
    }


    return render(request, "missed_refills.html", context)



